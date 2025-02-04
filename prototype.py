from distutils.command.build import build
from pdf2docx import parse
import docx
from simplify_docx import simplify
import json
import spacy
import os
from os.path import exists
import sqlite3
from sqlite3 import Error
import openpyxl
from pathlib import Path
from prompt_toolkit import prompt
from prompt_toolkit.history import FileHistory
from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
from prompt_toolkit.completion import WordCompleter

def get_connection(db_file):
    """ create a database connection to a SQLite database """
    return sqlite3.connect(db_file)

def initialize_database():
    xlsx_file = Path(".", "index.xlsx")
    wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
    print(wb_obj.sheetnames)
    index = wb_obj['Sheet']
    headers = [cell.value.strip() for cell in index[1]]

    source_path = "Papers/"
    docs_path = "Docs/"
    json_path = "JSON/"
    ents_path = "Ents/"

    print(headers)
    files_to_process = []
    for row in index.iter_rows(min_row=2):  # Ignore Header Row
        row_dictionary = {
            key: cell.value for key, cell in zip(headers, row)
        }
        row_dictionary['paper_docx'] = os.path.join(docs_path, row_dictionary['paper_pdf'] + '.docx')
        row_dictionary['paper_json'] = os.path.join(json_path, row_dictionary['paper_pdf'] + '.json')
        row_dictionary['paper_entities'] = os.path.join(ents_path, row_dictionary['paper_pdf'] + '.json')
        row_dictionary['paper_pdf'] = os.path.join(source_path, row_dictionary['paper_pdf'])
        files_to_process.append(row_dictionary)

    conn = get_connection("data/test_db.sqlite")
    cur = conn.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS papers (
        paper_id INTEGER PRIMARY KEY,
        paper_name TEXT NOT NULL UNIQUE,
        paper_pdf TEXT NOT NULL,
        paper_docx TEXT NOT NULL,
        paper_json TEXT NOT NULL,
        paper_entities TEXT NOT NULL
    );""")

    cur.execute("""CREATE TABLE IF NOT EXISTS entities (
        entity_id INTEGER PRIMARY KEY,
        entity_name TEXT NOT NULL,
        entity_type TEXT NOT NULL,
        UNIQUE(entity_name, entity_type)
    );""")

    cur.execute("""CREATE TABLE IF NOT EXISTS papers_have_entities (
        entity_id INTEGER,
        paper_id INTEGER,
        count INTEGER DEFAULT 1,
        FOREIGN KEY(entity_id) REFERENCES entities(entity_id),
        FOREIGN KEY(paper_id) REFERENCES papers(paper_id),
        PRIMARY KEY(entity_id, paper_id)
    );""")

    nlp = spacy.load("en_core_web_sm")

    unprocessed = []
    processed = []
    for path in os.listdir(source_path):
        find_file = list(filter(lambda file: file['paper_pdf'] == os.path.join(source_path, path), files_to_process))
        if len(find_file) == 0:
            unprocessed.append(os.path.join(source_path, path))
        elif os.path.isfile(os.path.join(source_path, path)):
            file_dict = find_file[-1]
            pdf_file = file_dict['paper_pdf']
            docx_file = file_dict['paper_docx']
            if not exists(docx_file):
                # convert pdf to docx
                parse(pdf_file, docx_file)
            # read in a document 
            doc = docx.Document(docx_file)
            if not exists(file_dict['paper_json']):
                with open(file_dict['paper_json'],'w') as outputjson:
                    json.dump(simplify(doc,  {"special-characters-as-text":False} ),outputjson)
            fullTextArray = []
            for para in doc.paragraphs:
                fullTextArray.append(para.text)

            entities = {"entities":[]}        
            if not exists(file_dict['paper_entities']):
                full_doc = nlp(''.join(fullTextArray))
                with open(file_dict['paper_entities'],'w') as outputents:
                    for e in full_doc.ents:
                        entities["entities"].append({"text": e.text, "start_char": e.start_char, "end_char": e.end_char, "label": e.label_})
                    json.dump(entities,outputents)
            else:
                with open(file_dict['paper_entities'], 'r') as reading:
                    entities = json.load(reading)
            cur.execute("""INSERT INTO papers(paper_name,paper_pdf,paper_docx,paper_json,paper_entities)
                VALUES(:paper_name,:paper_pdf,:paper_docx,:paper_json,:paper_entities)
                ON CONFLICT(paper_name) DO UPDATE SET
                    paper_pdf=:paper_pdf,
                    paper_docx=:paper_docx,
                    paper_json=:paper_json,
                    paper_entities=:paper_entities
                    RETURNING paper_id""", file_dict)
            file_id = cur.fetchone()[0]
            for entity in entities['entities']:
                print(entity)
                cur.execute("""INSERT INTO entities(entity_name, entity_type)
                    VALUES(:text,:label)
                    ON CONFLICT(entity_name, entity_type) DO UPDATE SET entity_name=:text
                        RETURNING entity_id""", entity)
                entity_id = cur.fetchone()[0]
                cur.execute("""INSERT INTO papers_have_entities(entity_id, paper_id)
                    VALUES(:entity_id,:paper_id)
                    ON CONFLICT(entity_id, paper_id) DO UPDATE SET count=count+1""", {"entity_id": entity_id, "paper_id": file_id})           
            processed.append(file_dict['paper_pdf'])
        else:
            unprocessed.append(os.path.join(source_path, path))
    conn.commit()
    print(files_to_process)
    print(unprocessed)
    print(processed)

def build_and_cases(query_string_array, query_parts, entry_index):
    """@brief processes user's query string to build the SQL where clause

    Args:
        query_string_array (list): user query string split into a list
        query_parts (dict): a dict containing the different parts of the SQL query. The keys are 'select', 'from', 'where', and 'limit'
        entry_index (int): index of current word in the query string array

    Returns:
        (int): reutrns the index of the current word in the query string array after processing
    """
    current_index = entry_index
    # Add an opening bracket to the where
    query_parts['where'].append(" ( ")
    # Check if the query string contains the word 'person' and if so, add entity type 'PERSON' to the where clause
    if current_index < len(query_string_array) and query_string_array[current_index] == 'person':
            current_index+=1
            query_parts['where'].append("entities.entity_type='PERSON'")
    # Check if the query string contains the word 'organisation' and if so, add entity type 'ORG' to the where clause
    if current_index < len(query_string_array) and query_string_array[current_index] == 'organisation':
        current_index+=1
        query_parts['where'].append("entities.entity_type='ORG'")
    # Check if the query string contains the word 'work' and if so, add entity type 'WORK_OF_ART' to the where clause
    if current_index < len(query_string_array) and query_string_array[current_index] == 'work':
        current_index+=1
        query_parts['where'].append("entities.entity_type='WORK_OF_ART'")
    # Add an AND between entity types and entity names
    query_parts['where'].append(" AND ")
    # Add the entity name to the where clause
    value_string = []
    while current_index < len(query_string_array) and query_string_array[current_index] != 'and' and query_string_array[current_index] != 'or':
        value_string.append(query_string_array[current_index])
        current_index+=1
    query_parts['where'].append(f"entities.entity_name = '{' '.join(value_string)}'")
    # Add a closing bracket to the where clause
    query_parts['where'].append(" ) ")
    # Return the index of the current word in the query string array after processing
    return current_index

def build_query(input_values):
    """@brief processes user's query string to build the SQL query

    Args:
        input_values (string): query string entered by the user

    Returns:
        (string): generated SQL query string
    """
    # Split the user entered query string into a list
    query_string_array = input_values.split(' ')
    # Initialize the different parts of the SQL query
    query_parts = {
        "select": [],
        "from": [],
        "where": [],
        "limit": []
    }
    # Initialize the SQL query string and the current index
    query_string = ""
    current_index = 0
    # check if the query string start with the word 'get' and if so, start building the SQL query
    if current_index < len(query_string_array) and query_string_array[current_index] == 'get':
        current_index+=1
        # Check if the query word after 'get' is 'one' and if so, limit the query to one result
        if current_index < len(query_string_array) and query_string_array[current_index] == 'one':
            current_index+=1
            query_parts['limit'].append(' Limit 1')
        # Check if the query word after 'get' is 'all' and if so, query all results
        elif current_index < len(query_string_array) and query_string_array[current_index] == 'all':
            current_index+=1

        # Check if the query word after 'one' or 'all' is 'papers' and if so, select all columns from the papers table
        if current_index < len(query_string_array) and query_string_array[current_index] == 'papers':
            current_index+=1
            query_parts['select'].append('*')
            query_parts['from'].append('papers')
        # Check if the query word after 'papers' is 'that' and if so, start building the where clause
        if current_index < len(query_string_array) and query_string_array[current_index] == 'that':
            current_index+=1
            # The query string must contain the word 'mention' to build the where clause
            if current_index < len(query_string_array) and query_string_array[current_index] == 'mention':
                current_index+=1
                # Add the inner join statements for the paper table ,papers_have_entities table and entities table
                query_parts['from'].append('INNER JOIN papers_have_entities ON papers_have_entities.paper_id = papers.paper_id')
                query_parts['from'].append('INNER JOIN entities ON papers_have_entities.entity_id = entities.entity_id')
                # call the build_and_cases function repeatedly to build the where clause
                while current_index < len(query_string_array):
                    current_index = build_and_cases(query_string_array, query_parts, current_index)
                    # Add an AND or OR between the different entity cases
                    if current_index < len(query_string_array):
                        if query_string_array[current_index] == 'and':
                            current_index += 1
                            query_parts['where'].append(" AND ")
                        elif query_string_array[current_index] == 'or':
                            current_index += 1
                            query_parts['where'].append(" OR ")
            # Generate the SQL query string based on the different parts of the query parts dict
            query_string = "SELECT " + ' '.join(query_parts['select']) + " FROM " + ' '.join(query_parts['from']) + " WHERE " + ' '.join(query_parts['where']) + ' ' + ' '.join(query_parts['limit'])
     
        else:
            # If the query string does not contain the word 'that' then the query do not have a where clause
            query_string = "SELECT " + ' '.join(query_parts['select']) + " FROM " + ' '.join(query_parts['from']) + ' '.join(query_parts['limit'])
    # Return the generated SQL query
    return query_string

def run_cli():
    # Create a WordCompleter object to suggest words to the user
    QuestionWorkCompleter = WordCompleter(['get', 'one', 'all', 'papers', 'that', 'mention', 'person', 'organisation', 'work', 'and', 'or'],
                                 ignore_case=True)
    # Create a connection to the database
    conn = get_connection("data/test_db.sqlite")
    cur = conn.cursor()

    # Create a loop to allow the user to enter queries until they enter 'q'
    user_input = None
    while user_input != 'q':
        # Prompt the user to enter a query string and store the history of the queries in the history file
        user_input = prompt('>',
                            history=FileHistory('historyPSD.txt'),
                            auto_suggest=AutoSuggestFromHistory(),
                            completer=QuestionWorkCompleter,)
        print(user_input)
        cur.execute(build_query(user_input))
        results = cur.fetchall()
        for result in results:
            print(result)

if __name__ == '__main__':
    initialize_database()
    run_cli()
